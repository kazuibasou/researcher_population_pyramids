import pickle
import time
import numpy as np
from sklearn.naive_bayes import ComplementNB, MultinomialNB
from sklearn.metrics import average_precision_score, roc_auc_score, f1_score, roc_curve
import openpyxl
import nomquamgender as nqg
import math
from sklearn.utils.class_weight import compute_sample_weight
import pandas as pd

def tuning_alpha_in_orbis_data_for_ComplementNB(country_lst, include_middle_name, include_last_name, clearning_special_chars):

    # alpha_lst = ([0.001 * i for i in range(1, 10)]
    #              + [0.01 * i for i in range(1, 10)]
    #              + [0.1 * i for i in range(1, 10)]
    #              + [i for i in range(1, 11)])
    alpha_lst = [0.001, 0.01, 0.1, 1, 10]

    gender_inference_accuracy = {}

    for country in country_lst:

        gender_inference_accuracy[country] = {}

        f_name = ("./data/" + str(country).lower() + "_train_val_test_sets"
                  + "_middle_name_" + str(include_middle_name)
                  + "_last_name_" + str(include_last_name)
                  + "_special_chars_" + str(clearning_special_chars)
                  + ".pkl")
        with open(f_name, 'rb') as f:
            c_data = pickle.load(f)
            X_train, y_train, X_val, y_val = c_data[0], c_data[1], c_data[2], c_data[3]

        sample_weight = compute_sample_weight(class_weight="balanced", y=y_train)

        for alpha in alpha_lst:
            train_t_s = time.time()
            clf = ComplementNB(alpha=alpha, force_alpha=True)
            clf.fit(X_train, y_train, sample_weight=sample_weight)
            train_t_e = time.time()

            test_t_s = time.time()
            y_pred_prob = clf.predict_proba(X_val)
            test_t_e = time.time()

            pr_auc = average_precision_score(y_val, y_pred_prob[:, 1])
            roc_auc = roc_auc_score(y_val, y_pred_prob[:, 1])

            fpr, tpr, thresholds = roc_curve(y_val, y_pred_prob[:, 1])

            prob_threshold = float('inf')
            for i in range(0, len(thresholds)):
                if tpr[i] >= 0.9:
                    prob_threshold = thresholds[i]
                    break

            y_, y_pred_ = [], []
            for i in range(0, len(y_val)):
                if y_pred_prob[i, 0] > y_pred_prob[i, 1]:
                    g, prob = 0, y_pred_prob[i, 0]
                elif y_pred_prob[i, 0] < y_pred_prob[i, 1]:
                    g, prob = 1, y_pred_prob[i, 1]
                else:
                    g, prob = -1, 0

                if prob >= prob_threshold:
                    y_.append(y_val[i])
                    y_pred_.append(g)

            # f1 = f1_score(np.array(y_, dtype=int), np.array(y_pred_, dtype=int), average='micro')

            gender_inference_accuracy[country][alpha] = {
                "PR AUC": pr_auc,
                "ROC AUC": roc_auc,
                "Threshold": prob_threshold,
                "N_val": len(y_val),
                "N_train": X_train.shape[0],
                "Gender assignment rate": float(len(y_)) / len(y_val),
                "Training time in seconds": train_t_e - train_t_s,
                "Testing time in seconds": test_t_e - test_t_s,
            }

            print("country", country, "alpha", alpha, "measure", gender_inference_accuracy[country][alpha])

    wb = openpyxl.Workbook()
    wb.create_sheet(index=0, title="ComplementNB")
    ws = wb["ComplementNB"]

    for i in range(0, len(alpha_lst)):
        ws.cell(1, 4 + 4 * i).value = "alpha = " + str(alpha_lst[i])
    ws.cell(1, 4 + 4 * len(alpha_lst)).value = "Best alpha"

    measure_lst = ['PR AUC', 'ROC AUC', 'Threshold', 'Gender assignment rate']
    ws.cell(2, 1).value = "Country"
    ws.cell(2, 2).value = "N_train"
    ws.cell(2, 3).value = "N_val"
    for i in range(0, len(alpha_lst)):
        for j in range(0, len(measure_lst)):
            ws.cell(2, 4 + 4 * i + j).value = measure_lst[j]
    for j in range(0, len(measure_lst)):
        ws.cell(2, 4 + 4 * len(alpha_lst) + j + 1).value = measure_lst[j]

    nb_best_alpha = {}
    for i in range(0, len(country_lst)):
        country = country_lst[i]

        ws.cell(3 + i, 1).value = str(country)

        if len(gender_inference_accuracy[country]) == 0:
            continue

        N_train, N_val = gender_inference_accuracy[country][alpha_lst[0]]['N_train'], \
            gender_inference_accuracy[country][alpha_lst[0]]['N_val']
        ws.cell(3 + i, 2).value = str(N_train)
        ws.cell(3 + i, 3).value = str(N_val)

        best_alpha, highest_auc = -1, 0
        for j in range(0, len(alpha_lst)):
            for k in range(0, len(measure_lst)):
                ws.cell(3 + i, 4 + 4 * j + k).value = str(
                    gender_inference_accuracy[country][alpha_lst[j]][measure_lst[k]])
                if str(measure_lst[k]) == 'ROC AUC' and highest_auc < \
                        gender_inference_accuracy[country][alpha_lst[j]][measure_lst[k]]:
                    highest_auc = gender_inference_accuracy[country][alpha_lst[j]][measure_lst[k]]
                    best_alpha = alpha_lst[j]
        ws.cell(3 + i, 4 + 4 * len(alpha_lst)).value = str(best_alpha)
        nb_best_alpha[country] = float(best_alpha)
        for k in range(0, len(measure_lst)):
            ws.cell(3 + i, 4 + 4 * len(alpha_lst) + k + 1).value = str(
                gender_inference_accuracy[country][best_alpha][measure_lst[k]])

    f_name = ("./data/tuning_alpha_in_orbis_for_ComplementNB"
              + "_middle_name_" + str(include_middle_name)
              + "_last_name_" + str(include_last_name)
              + "_special_chars_" + str(clearning_special_chars) + ".xlsx")
    wb.save(f_name)

    f_name = "./data/complementnb_best_alpha.pickle"
    with open(f_name, "wb") as f:
        pickle.dump(nb_best_alpha, f)

    return

def tuning_probability_threshold_in_orbis_data_for_ComplementNB(country_lst, include_middle_name, include_last_name, clearning_special_chars):

    f_name = "./data/complementnb_best_alpha.pickle"
    with open(f_name, "rb") as f:
        nb_alpha = pickle.load(f)

    prob_thresholds = ([0.9 + 0.01 * i for i in range(0, 10)]
                       + [0.99 + 0.001 * i for i in range(0, 10)]
                       + [0.999 + 0.0001 * i for i in range(0, 10)])

    gender_inference_accuracy = {}

    for country in country_lst:

        gender_inference_accuracy[country] = {}

        f_name = ("./data/" + str(country).lower() + "_train_val_test_sets"
                  + "_middle_name_" + str(include_middle_name)
                  + "_last_name_" + str(include_last_name)
                  + "_special_chars_" + str(clearning_special_chars)
                  + ".pkl")
        with open(f_name, 'rb') as f:
            c_data = pickle.load(f)
            X_train, y_train, X_val, y_val = c_data[0], c_data[1], c_data[2], c_data[3]

        alpha = nb_alpha[country]
        sample_weight = compute_sample_weight(class_weight="balanced", y=y_train)

        train_t_s = time.time()
        clf = ComplementNB(alpha=alpha, force_alpha=True)
        clf.fit(X_train, y_train, sample_weight=sample_weight)
        train_t_e = time.time()

        test_t_s = time.time()
        y_pred_prob = clf.predict_proba(X_val)
        test_t_e = time.time()

        for prob_threshold in prob_thresholds:
            y_, y_pred_ = [], []
            for i in range(0, len(y_val)):
                if y_pred_prob[i, 0] > y_pred_prob[i, 1]:
                    g, prob = 0, y_pred_prob[i, 0]
                elif y_pred_prob[i, 0] < y_pred_prob[i, 1]:
                    g, prob = 1, y_pred_prob[i, 1]
                else:
                    g, prob = -1, 0

                if prob >= prob_threshold:
                    y_.append(y_val[i])
                    y_pred_.append(g)

            f1 = f1_score(np.array(y_, dtype=int), np.array(y_pred_, dtype=int), average='micro')

            gender_inference_accuracy[country][prob_threshold] = {
                "alpha": alpha,
                "F1 score": f1,
                "N_val": len(y_val),
                "N_train": X_train.shape[0],
                "Gender assignment rate": float(len(y_)) / len(y_val),
                "Training time in seconds": train_t_e - train_t_s,
                "Testing time in seconds": test_t_e - test_t_s,
            }

    wb = openpyxl.Workbook()
    wb.create_sheet(index=0, title="ComplementNB")
    ws = wb["ComplementNB"]

    for i in range(0, len(prob_thresholds)):
        ws.cell(1, 4 + 4 * i).value = "threshold = " + str(prob_thresholds[i])
    ws.cell(1, 4 + 4 * len(prob_thresholds)).value = "Best threshold"

    measure_lst = ['alpha', 'F1 score', 'Gender assignment rate']
    ws.cell(2, 1).value = "Country"
    ws.cell(2, 2).value = "N_train"
    ws.cell(2, 3).value = "N_val"
    for i in range(0, len(prob_thresholds)):
        for j in range(0, len(measure_lst)):
            ws.cell(2, 4 + 4 * i + j).value = measure_lst[j]
    for j in range(0, len(measure_lst)):
        ws.cell(2, 4 + 4 * len(prob_thresholds) + j + 1).value = measure_lst[j]

    nb_prob_threshold = {}
    for i in range(0, len(country_lst)):
        country = country_lst[i]

        ws.cell(3 + i, 1).value = str(country)

        if len(gender_inference_accuracy[country]) == 0:
            continue

        N_train, N_val = gender_inference_accuracy[country][prob_thresholds[0]]['N_train'], \
        gender_inference_accuracy[country][prob_thresholds[0]]['N_val']
        ws.cell(3 + i, 2).value = str(N_train)
        ws.cell(3 + i, 3).value = str(N_val)

        threshold_candidates, f1_threshold = [], 0.9
        for j in range(0, len(prob_thresholds)):
            for k in range(0, len(measure_lst)):
                ws.cell(3 + i, 4 + 4 * j + k).value \
                    = str(gender_inference_accuracy[country][prob_thresholds[j]][measure_lst[k]])
                if (str(measure_lst[k]) == 'F1 score'
                        and gender_inference_accuracy[country][prob_thresholds[j]][measure_lst[k]] >= f1_threshold):
                    threshold_candidates.append(prob_thresholds[j])
        if len(threshold_candidates) > 0:
            best_threshold = min(threshold_candidates)
            ws.cell(3 + i, 4 + 4 * len(prob_thresholds)).value = str(best_threshold)
            nb_prob_threshold[country] = float(best_threshold)
            for k in range(0, len(measure_lst)):
                ws.cell(3 + i, 4 + 4 * len(prob_thresholds) + k + 1).value \
                    = str(gender_inference_accuracy[country][best_threshold][measure_lst[k]])
        else:
            best_threshold = max(prob_thresholds)
            ws.cell(3 + i, 4 + 4 * len(prob_thresholds)).value = str(best_threshold)
            nb_prob_threshold[country] = float(best_threshold)
            for k in range(0, len(measure_lst)):
                ws.cell(3 + i, 4 + 4 * len(prob_thresholds) + k + 1).value \
                    = str(gender_inference_accuracy[country][max(prob_thresholds)][measure_lst[k]])

    f_name = ("./data/tuning_probability_threshold_in_orbis_for_ComplementNB"
              + "_middle_name_" + str(include_middle_name)
              + "_last_name_" + str(include_last_name)
              + "_special_chars_" + str(clearning_special_chars) + ".xlsx")
    wb.save(f_name)

    f_name = "./data/complementnb_prob_threshold_in_orbis.pickle"
    with open(f_name, "wb") as f:
        pickle.dump(nb_prob_threshold, f)

    return

def tuning_probability_threshold_in_orbis_data_for_nqg(country_lst, include_middle_name, include_last_name, clearning_special_chars):

    prob_thresholds = ([0.9 + 0.01 * i for i in range(0, 10)]
                       + [0.99 + 0.001 * i for i in range(0, 10)]
                       + [0.999 + 0.0001 * i for i in range(0, 10)])

    gender_inference_accuracy = {}

    for country in country_lst:

        gender_inference_accuracy[country] = {}

        f_name = ("./data/" + str(country).lower() + "_train_val_test_sets"
                  + "_middle_name_" + str(include_middle_name)
                  + "_last_name_" + str(include_last_name)
                  + "_special_chars_" + str(clearning_special_chars)
                  + ".pkl")
        with open(f_name, 'rb') as f:
            c_data = pickle.load(f)
            X_train, y_val, val_name_country_lst = c_data[0], c_data[3], c_data[7]

        test_names = []
        for i in range(0, len(y_val)):
            name = str(val_name_country_lst[i][0])
            test_names.append(name)

        model = nqg.NBGC()
        df = model.annotate(test_names, as_df=True)
        p_g = df['p(gf)'].to_list()

        y_pred_prob = np.zeros((len(y_val), 2))
        for i in range(0, len(p_g)):
            p = float(p_g[i])
            if math.isnan(p):
                p = 0.5
            y_pred_prob[i][0] = p
            y_pred_prob[i][1] = 1 - p

        for prob_threshold in prob_thresholds:
            y_, y_pred_ = [], []
            for i in range(0, len(y_val)):
                if y_pred_prob[i, 0] > y_pred_prob[i, 1]:
                    g, prob = 0, y_pred_prob[i, 0]
                elif y_pred_prob[i, 0] < y_pred_prob[i, 1]:
                    g, prob = 1, y_pred_prob[i, 1]
                else:
                    g, prob = -1, 0

                if prob >= prob_threshold:
                    y_.append(y_val[i])
                    y_pred_.append(g)

            f1 = f1_score(np.array(y_, dtype=int), np.array(y_pred_, dtype=int), average='micro')
            pr_auc = average_precision_score(y_val, y_pred_prob[:, 1])
            roc_auc = roc_auc_score(y_val, y_pred_prob[:, 1])

            gender_inference_accuracy[country][prob_threshold] = {
                "F1 score": f1,
                "N_val": len(y_val),
                "N_train": X_train.shape[0],
                "Gender assignment rate": float(len(y_)) / len(y_val),
                "PR AUC": pr_auc,
                "ROC AUC": roc_auc,
            }

    wb = openpyxl.Workbook()
    wb.create_sheet(index=0, title="nqg")
    ws = wb["nqg"]

    for i in range(0, len(prob_thresholds)):
        ws.cell(1, 4 + 4 * i).value = "threshold = " + str(prob_thresholds[i])
    ws.cell(1, 4 + 4 * len(prob_thresholds)).value = "Best threshold"

    measure_lst = ['PR AUC', 'ROC AUC', 'F1 score', 'Gender assignment rate']
    ws.cell(2, 1).value = "Country"
    ws.cell(2, 2).value = "N_train"
    ws.cell(2, 3).value = "N_val"
    for i in range(0, len(prob_thresholds)):
        for j in range(0, len(measure_lst)):
            ws.cell(2, 4 + 4 * i + j).value = measure_lst[j]
    for j in range(0, len(measure_lst)):
        ws.cell(2, 4 + 4 * len(prob_thresholds) + j + 1).value = measure_lst[j]

    nqg_prob_threshold = {}
    for i in range(0, len(country_lst)):
        country = country_lst[i]

        ws.cell(3 + i, 1).value = str(country)

        if len(gender_inference_accuracy[country]) == 0:
            continue

        N_train, N_val = gender_inference_accuracy[country][prob_thresholds[0]]['N_train'], \
        gender_inference_accuracy[country][prob_thresholds[0]]['N_val']
        ws.cell(3 + i, 2).value = str(N_train)
        ws.cell(3 + i, 3).value = str(N_val)

        threshold_candidates, f1_threshold = [], 0.9
        for j in range(0, len(prob_thresholds)):
            for k in range(0, len(measure_lst)):
                ws.cell(3 + i, 4 + 4 * j + k).value \
                    = str(gender_inference_accuracy[country][prob_thresholds[j]][measure_lst[k]])
                if (str(measure_lst[k]) == 'F1 score'
                        and gender_inference_accuracy[country][prob_thresholds[j]][measure_lst[k]] >= f1_threshold):
                    threshold_candidates.append(prob_thresholds[j])
        if len(threshold_candidates) > 0:
            best_threshold = min(threshold_candidates)
            ws.cell(3 + i, 4 + 4 * len(prob_thresholds)).value = str(best_threshold)
            nqg_prob_threshold[country] = float(best_threshold)
            for k in range(0, len(measure_lst)):
                ws.cell(3 + i, 4 + 4 * len(prob_thresholds) + k + 1).value \
                    = str(gender_inference_accuracy[country][best_threshold][measure_lst[k]])
        else:
            best_threshold = max(prob_thresholds)
            ws.cell(3 + i, 4 + 4 * len(prob_thresholds)).value = str(best_threshold)
            nqg_prob_threshold[country] = float(best_threshold)
            for k in range(0, len(measure_lst)):
                ws.cell(3 + i, 4 + 4 * len(prob_thresholds) + k + 1).value \
                    = str(gender_inference_accuracy[country][max(prob_thresholds)][measure_lst[k]])

    f_name = ("./data/tuning_probability_threshold_in_orbis_for_nqg.xlsx")
    wb.save(f_name)

    f_name = "./data/nqg_prob_threshold_in_orbis.pickle"
    with open(f_name, "wb") as f:
        pickle.dump(nqg_prob_threshold, f)

    return

def tuning_probability_threshold_in_wgnd_data_for_ComplementNB(country_lst, include_middle_name, include_last_name, clearning_special_chars):

    f_val_name = ("./data/wgnd_val_data.pkl")
    with open(f_val_name, mode="rb") as f:
        (wgnd_val_name_corpus, wgnd_y_val) = pickle.load(f)

    f_name = "./data/complementnb_best_alpha.pickle"
    with open(f_name, "rb") as f:
        nb_alpha = pickle.load(f)

    prob_thresholds = ([0.9 + 0.01 * i for i in range(0, 10)]
                       + [0.99 + 0.001 * i for i in range(0, 10)]
                       + [0.999 + 0.0001 * i for i in range(0, 10)])

    gender_inference_accuracy = {}

    for country in country_lst:

        gender_inference_accuracy[country] = {}

        f_name = ("./data/" + str(country).lower() + "_train_val_test_sets"
                  + "_middle_name_" + str(include_middle_name)
                  + "_last_name_" + str(include_last_name)
                  + "_special_chars_" + str(clearning_special_chars)
                  + ".pkl")
        with open(f_name, 'rb') as f:
            c_data = pickle.load(f)
            X_train, y_train = c_data[0], c_data[1]

        f_name = "./data/global_feature_transformers.pkl"
        with open(f_name, 'rb') as f:
            (count_vect, tfidf_transformer) = pickle.load(f)

        alpha = nb_alpha[country]

        if len(wgnd_y_val[country]) <= 0 or len(wgnd_val_name_corpus[country]) <= 0:
            continue

        X_val_counts = count_vect.transform(wgnd_val_name_corpus[country])
        X_val = tfidf_transformer.transform(X_val_counts)

        y_val = np.array(wgnd_y_val[country])

        print(country, X_train.shape, X_val.shape, y_train.shape, len(y_val))

        sample_weight = compute_sample_weight(class_weight="balanced", y=y_train)

        clf = ComplementNB(alpha=nb_alpha[country], force_alpha=True)
        clf.fit(X_train, y_train, sample_weight=sample_weight)

        y_pred_prob = clf.predict_proba(X_val)

        pr_auc = average_precision_score(wgnd_y_val[country], y_pred_prob[:, 1])
        roc_auc = roc_auc_score(wgnd_y_val[country], y_pred_prob[:, 1])

        for prob_threshold in prob_thresholds:
            y_, y_pred_ = [], []
            for i in range(0, len(y_val)):
                if y_pred_prob[i, 0] > y_pred_prob[i, 1]:
                    g, prob = 0, y_pred_prob[i, 0]
                elif y_pred_prob[i, 0] < y_pred_prob[i, 1]:
                    g, prob = 1, y_pred_prob[i, 1]
                else:
                    g, prob = -1, 0

                if prob >= prob_threshold:
                    y_.append(y_val[i])
                    y_pred_.append(g)

            f1 = f1_score(np.array(y_, dtype=int), np.array(y_pred_, dtype=int), average='micro')

            gender_inference_accuracy[country][prob_threshold] = {
                "alpha": alpha,
                "PR AUC": pr_auc,
                "ROC AUC": roc_auc,
                "F1 score": f1,
                "N_val": len(y_val),
                "N_train": X_train.shape[0],
                "Gender assignment rate": float(len(y_)) / len(y_val),
            }

    wb = openpyxl.Workbook()
    wb.create_sheet(index=0, title="ComplementNB")
    ws = wb["ComplementNB"]

    for i in range(0, len(prob_thresholds)):
        ws.cell(1, 4 + 4 * i).value = "threshold = " + str(prob_thresholds[i])
    ws.cell(1, 4 + 4 * len(prob_thresholds)).value = "Best threshold"

    measure_lst = ['alpha', "PR AUC", "ROC AUC", 'F1 score', "Gender assignment rate"]
    ws.cell(2, 1).value = "Country"
    ws.cell(2, 2).value = "N_train"
    ws.cell(2, 3).value = "N_val"
    for i in range(0, len(prob_thresholds)):
        for j in range(0, len(measure_lst)):
            ws.cell(2, 4 + 4 * i + j).value = measure_lst[j]
    for j in range(0, len(measure_lst)):
        ws.cell(2, 4 + 4 * len(prob_thresholds) + j + 1).value = measure_lst[j]

    nb_prob_threshold = {}
    for i in range(0, len(country_lst)):
        country = country_lst[i]

        ws.cell(3 + i, 1).value = str(country)

        if len(gender_inference_accuracy[country]) == 0:
            continue

        N_train, N_val = gender_inference_accuracy[country][prob_thresholds[0]]['N_train'], \
            gender_inference_accuracy[country][prob_thresholds[0]]['N_val']
        ws.cell(3 + i, 2).value = str(N_train)
        ws.cell(3 + i, 3).value = str(N_val)

        threshold_candidates, f1_threshold = [], 0.9
        for j in range(0, len(prob_thresholds)):
            for k in range(0, len(measure_lst)):
                ws.cell(3 + i, 4 + 4 * j + k).value \
                    = str(gender_inference_accuracy[country][prob_thresholds[j]][measure_lst[k]])
                if (str(measure_lst[k]) == 'F1 score'
                        and gender_inference_accuracy[country][prob_thresholds[j]][measure_lst[k]] >= f1_threshold):
                    threshold_candidates.append(prob_thresholds[j])
        if len(threshold_candidates) > 0:
            best_threshold = min(threshold_candidates)
            ws.cell(3 + i, 4 + 4 * len(prob_thresholds)).value = str(best_threshold)
            nb_prob_threshold[country] = float(best_threshold)
            for k in range(0, len(measure_lst)):
                ws.cell(3 + i, 4 + 4 * len(prob_thresholds) + k + 1).value \
                    = str(gender_inference_accuracy[country][best_threshold][measure_lst[k]])
        else:
            best_threshold = max(prob_thresholds)
            ws.cell(3 + i, 4 + 4 * len(prob_thresholds)).value = str(best_threshold)
            nb_prob_threshold[country] = float(best_threshold)
            for k in range(0, len(measure_lst)):
                ws.cell(3 + i, 4 + 4 * len(prob_thresholds) + k + 1).value \
                    = str(gender_inference_accuracy[country][max(prob_thresholds)][measure_lst[k]])

    f_name = ("./data/tuning_probability_threshold_in_wgnd_for_ComplementNB"
              + "_middle_name_" + str(include_middle_name)
              + "_last_name_" + str(include_last_name)
              + "_special_chars_" + str(clearning_special_chars) + ".xlsx")
    wb.save(f_name)

    f_name = "./data/complementnb_prob_threshold_in_wgnd.pickle"
    with open(f_name, "wb") as f:
        pickle.dump(nb_prob_threshold, f)

    return

def tuning_probability_threshold_in_wgnd_data_for_nqg(country_lst):

    wb = openpyxl.load_workbook("./data/wgnd_val_data_for_api.xlsx")
    ws = wb["Sheet1"]

    y_val_ = {country: [] for country in country_lst}
    y_pred_prob_ = {country: [] for country in country_lst}
    read = True
    i = 0

    while read:
        x1, x2, x3 = str(ws.cell(i + 2, 1).value), str(ws.cell(i + 2, 2).value), str(ws.cell(i + 2, 3).value)
        i += 1

        if x1 == 'None' and x2 == 'None' and x3 == 'None':
            break

        first_name = str(x1)
        country = str(x2)
        gender = int(x3)

        if country not in country_lst:
            continue

        model = nqg.NBGC()
        p_g = float(model.annotate(first_name)[0][4])
        if math.isnan(p_g):
            p_g = 0.5

        y_val_[country].append(gender)
        y_pred_prob_[country].append([p_g, 1 - p_g])

    prob_thresholds = ([0.9 + 0.01 * i for i in range(0, 10)]
                       + [0.99 + 0.001 * i for i in range(0, 10)]
                       + [0.999 + 0.0001 * i for i in range(0, 10)])

    gender_inference_accuracy = {}

    for country in country_lst:

        gender_inference_accuracy[country] = {}

        y_val = np.array(y_val_[country])
        y_pred_prob = np.array(y_pred_prob_[country])

        if len(y_val) <= 0 or len(y_pred_prob) <= 0:
            continue

        pr_auc = average_precision_score(y_val, y_pred_prob[:, 1])
        roc_auc = roc_auc_score(y_val, y_pred_prob[:, 1])

        for prob_threshold in prob_thresholds:
            y_, y_pred_ = [], []
            for i in range(0, len(y_val)):
                if y_pred_prob[i, 0] > y_pred_prob[i, 1]:
                    g, prob = 0, y_pred_prob[i, 0]
                elif y_pred_prob[i, 0] < y_pred_prob[i, 1]:
                    g, prob = 1, y_pred_prob[i, 1]
                else:
                    g, prob = -1, 0

                if prob >= prob_threshold:
                    y_.append(y_val[i])
                    y_pred_.append(g)

            f1 = f1_score(np.array(y_, dtype=int), np.array(y_pred_, dtype=int), average='micro')

            gender_inference_accuracy[country][prob_threshold] = {
                "PR AUC": pr_auc,
                "ROC AUC": roc_auc,
                "F1 score": f1,
                "N_val": len(y_val),
                "Gender assignment rate": float(len(y_)) / len(y_val),
            }

    wb = openpyxl.Workbook()
    wb.create_sheet(index=0, title="nqg")
    ws = wb["nqg"]

    for i in range(0, len(prob_thresholds)):
        ws.cell(1, 4 + 4 * i).value = "threshold = " + str(prob_thresholds[i])
    ws.cell(1, 4 + 4 * len(prob_thresholds)).value = "Best threshold"

    measure_lst = ['PR AUC', 'ROC AUC', 'F1 score', 'Gender assignment rate']
    ws.cell(2, 1).value = "Country"
    ws.cell(2, 2).value = "N_train"
    ws.cell(2, 3).value = "N_val"
    for i in range(0, len(prob_thresholds)):
        for j in range(0, len(measure_lst)):
            ws.cell(2, 4 + 4 * i + j).value = measure_lst[j]
    for j in range(0, len(measure_lst)):
        ws.cell(2, 4 + 4 * len(prob_thresholds) + j + 1).value = measure_lst[j]

    nqg_prob_threshold = {}
    for i in range(0, len(country_lst)):
        country = country_lst[i]

        ws.cell(3 + i, 1).value = str(country)

        if len(gender_inference_accuracy[country]) == 0:
            continue

        N_val = gender_inference_accuracy[country][prob_thresholds[0]]['N_val']
        ws.cell(3 + i, 3).value = str(N_val)

        threshold_candidates, f1_threshold = [], 0.9
        for j in range(0, len(prob_thresholds)):
            for k in range(0, len(measure_lst)):
                ws.cell(3 + i, 4 + 4 * j + k).value \
                    = str(gender_inference_accuracy[country][prob_thresholds[j]][measure_lst[k]])
                if (str(measure_lst[k]) == 'F1 score'
                        and gender_inference_accuracy[country][prob_thresholds[j]][measure_lst[k]] >= f1_threshold):
                    threshold_candidates.append(prob_thresholds[j])
        if len(threshold_candidates) > 0:
            best_threshold = min(threshold_candidates)
            ws.cell(3 + i, 4 + 4 * len(prob_thresholds)).value = str(best_threshold)
            nqg_prob_threshold[country] = float(best_threshold)
            for k in range(0, len(measure_lst)):
                ws.cell(3 + i, 4 + 4 * len(prob_thresholds) + k + 1).value \
                    = str(gender_inference_accuracy[country][best_threshold][measure_lst[k]])
        else:
            best_threshold = max(prob_thresholds)
            ws.cell(3 + i, 4 + 4 * len(prob_thresholds)).value = str(best_threshold)
            nqg_prob_threshold[country] = float(best_threshold)
            for k in range(0, len(measure_lst)):
                ws.cell(3 + i, 4 + 4 * len(prob_thresholds) + k + 1).value \
                    = str(gender_inference_accuracy[country][max(prob_thresholds)][measure_lst[k]])

    f_name = ("./data/tuning_probability_threshold_in_wgnd_for_nqg.xlsx")
    wb.save(f_name)

    f_name = "./data/nqg_prob_threshold_in_wgnd.pickle"
    with open(f_name, "wb") as f:
        pickle.dump(nqg_prob_threshold, f)

    return

def set_cnb_params(country_lst):

    f_name = "./data/complementnb_best_alpha.pickle"
    with open(f_name, "rb") as f:
        nb_alpha = pickle.load(f)

    f_name = "./data/complementnb_prob_threshold_in_orbis.pickle"
    with open(f_name, "rb") as f:
        nb_prob_threshold_orbis = pickle.load(f)

    f_name = "./data/complementnb_prob_threshold_in_wgnd.pickle"
    with open(f_name, "rb") as f:
        nb_prob_threshold_wgnd = pickle.load(f)

    nb_prob_threshold = {}
    for country in country_lst:
        if country not in nb_prob_threshold_orbis or country not in nb_prob_threshold_wgnd:
            continue
        p1 = nb_prob_threshold_orbis[country]
        p2 = nb_prob_threshold_wgnd[country]
        nb_prob_threshold[country] = max(p1, p2)

    return nb_alpha, nb_prob_threshold

def test_ComplementNB_in_orbis_data(country_lst, include_middle_name, include_last_name, clearning_special_chars):

    nb_alpha, nb_prob_threshold = set_cnb_params(country_lst)

    test_results = []

    for country in country_lst:

        if country not in nb_alpha or country not in nb_prob_threshold:
            test_results.append([
                country,
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ])
            print(country, "Not performed test.")
            continue

        f_name = ("./data/" + str(country).lower() + "_train_val_test_sets"
                  + "_middle_name_" + str(include_middle_name)
                  + "_last_name_" + str(include_last_name)
                  + "_special_chars_" + str(clearning_special_chars)
                  + ".pkl")
        with open(f_name, 'rb') as f:
            c_data = pickle.load(f)
            X_train, y_train, X_test, y_test = c_data[0], c_data[1], c_data[4], c_data[5]

        alpha = nb_alpha[country]
        sample_weight = compute_sample_weight(class_weight="balanced", y=y_train)

        train_t_s = time.time()
        clf = ComplementNB(alpha=alpha, force_alpha=True)
        clf.fit(X_train, y_train, sample_weight=sample_weight)
        train_t_e = time.time()

        test_t_s = time.time()
        y_pred_prob = clf.predict_proba(X_test)
        test_t_e = time.time()

        pr_auc = average_precision_score(y_test, y_pred_prob[:, 1])
        roc_auc = roc_auc_score(y_test, y_pred_prob[:, 1])

        prob_threshold = nb_prob_threshold[country]

        y_, y_pred_ = [], []
        for i in range(0, len(y_test)):
            if y_pred_prob[i, 0] > y_pred_prob[i, 1]:
                g, prob = 0, y_pred_prob[i, 0]
            elif y_pred_prob[i, 0] < y_pred_prob[i, 1]:
                g, prob = 1, y_pred_prob[i, 1]
            else:
                g, prob = -1, 0

            if prob >= prob_threshold:
                y_.append(y_test[i])
                y_pred_.append(g)

        f1 = f1_score(np.array(y_, dtype=int), np.array(y_pred_, dtype=int), average='micro')

        test_results.append([
            country,
            alpha,
            prob_threshold,
            pr_auc,
            roc_auc,
            f1,
            len(y_test),
            X_train.shape[0],
            float(len(y_)) / len(y_test),
            train_t_e - train_t_s,
            test_t_e - test_t_s,
        ])
        print(country,
            alpha,
            prob_threshold,
            pr_auc,
            roc_auc,
            f1,
            len(y_test),
            X_train.shape[0],
            float(len(y_)) / len(y_test),
            train_t_e - train_t_s,
            test_t_e - test_t_s,
        )

    df = pd.DataFrame(test_results,
                      columns=['country', 'alpha', 'prob_threshold',
                               'PR AUC', 'ROC AUC',
                               'f1', 'N_test', 'N_train', 'gender assignment rate',
                               'training time', 'test time'])
    df.to_excel('./data/orbis_test_results_ComplementNB.xlsx', index=False)

    return

def test_ComplementNB_in_wgnd_data(country_lst, include_middle_name, include_last_name, clearning_special_chars):

    nb_alpha, nb_prob_threshold = set_cnb_params(country_lst)

    test_results = []

    f_val_name = ("./data/wgnd_test_data.pkl")
    with open(f_val_name, mode="rb") as f:
        (wgnd_test_name_corpus, wgnd_y_test) = pickle.load(f)

    for country in country_lst:

        if country not in nb_alpha or country not in nb_prob_threshold:
            test_results.append([
                country,
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ])
            print(country, "Not performed test.")
            continue

        f_name = ("./data/" + str(country).lower() + "_train_val_test_sets"
                  + "_middle_name_" + str(include_middle_name)
                  + "_last_name_" + str(include_last_name)
                  + "_special_chars_" + str(clearning_special_chars)
                  + ".pkl")
        with open(f_name, 'rb') as f:
            c_data = pickle.load(f)
            X_train, y_train = c_data[0], c_data[1]

        f_name = "./data/global_feature_transformers.pkl"
        with open(f_name, 'rb') as f:
            (count_vect, tfidf_transformer) = pickle.load(f)

        alpha = nb_alpha[country]

        if len(wgnd_y_test[country]) <= 0 or len(wgnd_test_name_corpus[country]) <= 0:
            continue

        X_val_counts = count_vect.transform(wgnd_test_name_corpus[country])
        X_val = tfidf_transformer.transform(X_val_counts)

        y_test = np.array(wgnd_y_test[country])

        sample_weight = compute_sample_weight(class_weight="balanced", y=y_train)

        clf = ComplementNB(alpha=nb_alpha[country], force_alpha=True)
        clf.fit(X_train, y_train, sample_weight=sample_weight)

        y_pred_prob = clf.predict_proba(X_val)

        pr_auc = average_precision_score(y_test, y_pred_prob[:, 1])
        roc_auc = roc_auc_score(y_test, y_pred_prob[:, 1])

        prob_threshold = nb_prob_threshold[country]

        y_, y_pred_ = [], []
        for i in range(0, len(y_test)):
            if y_pred_prob[i, 0] > y_pred_prob[i, 1]:
                g, prob = 0, y_pred_prob[i, 0]
            elif y_pred_prob[i, 0] < y_pred_prob[i, 1]:
                g, prob = 1, y_pred_prob[i, 1]
            else:
                g, prob = -1, 0

            if prob >= prob_threshold:
                y_.append(y_test[i])
                y_pred_.append(g)

        f1 = f1_score(np.array(y_, dtype=int), np.array(y_pred_, dtype=int), average='micro')

        test_results.append([
            country,
            alpha,
            prob_threshold,
            pr_auc,
            roc_auc,
            f1,
            len(y_test),
            X_train.shape[0],
            float(len(y_)) / len(y_test),
        ])
        print(country,
              alpha,
              prob_threshold,
              pr_auc,
              roc_auc,
              f1,
              len(y_test),
              X_train.shape[0],
              float(len(y_)) / len(y_test),
              )

    df = pd.DataFrame(test_results,
                      columns=['country', 'alpha', 'prob_threshold',
                               'PR AUC', 'ROC AUC', 'f1', 'N_test', 'N_train',
                               'gender assignment rate'])
    df.to_excel('./data/wgnd_test_results_ComplementNB.xlsx', index=False)

    return

def set_nqg_param(country_lst):

    f_name = "./data/nqg_prob_threshold_in_orbis.pickle"
    with open(f_name, "rb") as f:
        nqg_prob_threshold_orbis = pickle.load(f)

    f_name = "./data/nqg_prob_threshold_in_wgnd.pickle"
    with open(f_name, "rb") as f:
        nqg_prob_threshold_wgnd = pickle.load(f)

    nqg_prob_threshold = {}
    for country in country_lst:
        if country not in nqg_prob_threshold_orbis or country not in nqg_prob_threshold_wgnd:
            continue
        p1 = nqg_prob_threshold_orbis[country]
        p2 = nqg_prob_threshold_wgnd[country]
        nqg_prob_threshold[country] = max(p1, p2)

    print(nqg_prob_threshold)

    return nqg_prob_threshold

def test_nqg_in_orbis_data(country_lst, include_middle_name, include_last_name, clearning_special_chars):

    nqg_prob_threshold = set_nqg_param(country_lst)
    test_results = []

    for country in country_lst:

        if country not in nqg_prob_threshold:
            test_results.append([
                country,
                "",
                "",
                "",
                "",
                "",
                "",
            ])
            print(country, "Not performed test.")
            continue

        f_name = ("./data/" + str(country).lower() + "_train_val_test_sets"
                  + "_middle_name_" + str(include_middle_name)
                  + "_last_name_" + str(include_last_name)
                  + "_special_chars_" + str(clearning_special_chars)
                  + ".pkl")
        with open(f_name, 'rb') as f:
            c_data = pickle.load(f)
            y_test, test_name_country_lst = c_data[5], c_data[6]

        test_names = []
        for i in range(0, len(y_test)):
            name = str(test_name_country_lst[i][0])
            test_names.append(name)

        model = nqg.NBGC()
        df = model.annotate(test_names, as_df=True)
        p_g = df['p(gf)'].to_list()

        y_pred_prob = np.zeros((len(y_test), 2))
        for i in range(0, len(p_g)):
            p = float(p_g[i])
            if math.isnan(p):
                p = 0.5
            y_pred_prob[i][0] = p
            y_pred_prob[i][1] = 1 - p

        prob_threshold = nqg_prob_threshold[country]

        y_, y_pred_ = [], []
        for i in range(0, len(y_test)):
            if y_pred_prob[i, 0] > y_pred_prob[i, 1]:
                g, prob = 0, y_pred_prob[i, 0]
            elif y_pred_prob[i, 0] < y_pred_prob[i, 1]:
                g, prob = 1, y_pred_prob[i, 1]
            else:
                g, prob = -1, 0

            if prob >= prob_threshold:
                y_.append(y_test[i])
                y_pred_.append(g)

        f1 = f1_score(np.array(y_, dtype=int), np.array(y_pred_, dtype=int), average='micro')
        pr_auc = average_precision_score(y_test, y_pred_prob[:, 1])
        roc_auc = roc_auc_score(y_test, y_pred_prob[:, 1])

        test_results.append([
            country,
            prob_threshold,
            pr_auc,
            roc_auc,
            f1,
            len(y_test),
            float(len(y_)) / len(y_test),
        ])
        print([
            country,
            prob_threshold,
            pr_auc,
            roc_auc,
            f1,
            len(y_test),
            float(len(y_)) / len(y_test),
        ])

    df = pd.DataFrame(test_results,
                      columns=['country', 'prob_threshold',
                               'PR AUC', 'ROC AUC', 'f1', 'N_test',
                               'gender assignment rate'])
    df.to_excel('./data/orbis_test_results_nqg.xlsx', index=False)

    return

def test_nqg_in_wgnd_data(country_lst):

    wb = openpyxl.load_workbook("./data/wgnd_test_data_for_api.xlsx")
    ws = wb["Sheet1"]

    y_test_ = {country: [] for country in country_lst}
    y_pred_prob_ = {country: [] for country in country_lst}
    read = True
    i = 0

    while read:
        x1, x2, x3 = str(ws.cell(i + 2, 1).value), str(ws.cell(i + 2, 2).value), str(ws.cell(i + 2, 3).value)
        i += 1

        if x1 == 'None' and x2 == 'None' and x3 == 'None':
            break

        first_name = str(x1)
        country = str(x2)
        gender = int(x3)

        if country not in country_lst:
            continue

        model = nqg.NBGC()
        p_g = float(model.annotate(first_name)[0][4])
        if math.isnan(p_g):
            p_g = 0.5

        y_test_[country].append(gender)
        y_pred_prob_[country].append([p_g, 1 - p_g])

    nqg_prob_threshold = set_nqg_param(country_lst)
    test_results = []

    for country in country_lst:

        if country not in nqg_prob_threshold:
            test_results.append([
                country,
                "",
                "",
                "",
                "",
                "",
                "",
            ])
            print(country, "Not performed test.")
            continue

        y_test = np.array(y_test_[country])
        y_pred_prob = np.array(y_pred_prob_[country])

        if len(y_test) <= 0 or len(y_pred_prob) <= 0:
            continue

        prob_threshold = nqg_prob_threshold[country]

        y_, y_pred_ = [], []
        for i in range(0, len(y_test)):
            if y_pred_prob[i, 0] > y_pred_prob[i, 1]:
                g, prob = 0, y_pred_prob[i, 0]
            elif y_pred_prob[i, 0] < y_pred_prob[i, 1]:
                g, prob = 1, y_pred_prob[i, 1]
            else:
                g, prob = -1, 0

            if prob >= prob_threshold:
                y_.append(y_test[i])
                y_pred_.append(g)

        f1 = f1_score(np.array(y_, dtype=int), np.array(y_pred_, dtype=int), average='micro')
        pr_auc = average_precision_score(y_test, y_pred_prob[:, 1])
        roc_auc = roc_auc_score(y_test, y_pred_prob[:, 1])

        test_results.append([
            country,
            prob_threshold,
            pr_auc,
            roc_auc,
            f1,
            len(y_test),
            float(len(y_)) / len(y_test),
        ])
        print([
            country,
            prob_threshold,
            pr_auc,
            roc_auc,
            f1,
            len(y_test),
            float(len(y_)) / len(y_test),
        ])

    df = pd.DataFrame(test_results,
                      columns=['country', 'prob_threshold',
                               'PR AUC', 'ROC AUC', 'f1', 'N_test',
                               'gender assignment rate'])
    df.to_excel('./data/wgnd_test_results_nqg.xlsx', index=False)

    return